from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from pypdf import PdfReader
import pdfplumber
from docx import Document


SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".xlsx",
    ".xls",
    ".csv",
    ".txt",
    ".md",
    ".json",
    ".xml",
    ".yaml",
    ".yml",
}


def extract_text(file_path: Path) -> str:
    """Extrae texto de documentos PDF, DOCX, Excel, CSV y texto plano."""
    return extract_document_content(file_path)["text"]


def extract_document_content(file_path: Path) -> Dict[str, Any]:
    """Extrae contenido de un documento y devuelve texto junto con contexto adicional."""
    extension = file_path.suffix.lower()

    if extension == ".pdf":
        return extract_pdf_content(file_path)
    if extension == ".docx":
        return extract_docx_content(file_path)
    if extension in {".xlsx", ".xls"}:
        return extract_excel_content(file_path)
    if extension == ".csv":
        return {"text": extract_csv_text(file_path), "location_hint": "contenido principal"}
    if extension in {".txt", ".md", ".json", ".xml", ".yaml", ".yml"}:
        return {"text": read_plain_text(file_path), "location_hint": "contenido principal"}

    return {"text": "", "location_hint": "contenido principal"}


def extract_pdf_content(file_path: Path) -> Dict[str, Any]:
    """Extrae texto de PDF y conserva información de páginas para la ubicación."""
    page_entries: List[Dict[str, Any]] = []
    try:
        reader = PdfReader(str(file_path))
        for page_number, page in enumerate(reader.pages, start=1):
            extracted = (page.extract_text() or "").strip()
            if extracted:
                page_entries.append({"page_number": page_number, "text": extracted})
        if page_entries:
            text = "\n".join([p["text"] for p in page_entries if p["text"]]).strip()
            return {
                "text": text,
                "pages": page_entries,
                "location_hint": f"página 1-{len(page_entries)}",
            }
    except Exception:
        pass

    try:
        with pdfplumber.open(str(file_path)) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                extracted = (page.extract_text() or "").strip()
                if extracted:
                    page_entries.append({"page_number": page_number, "text": extracted})
            text = "\n".join([p["text"] for p in page_entries if p["text"]]).strip()
            return {
                "text": text,
                "pages": page_entries,
                "location_hint": f"página 1-{len(page_entries)}",
            }
    except Exception:
        return {"text": "", "pages": [], "location_hint": "sin contenido"}


def extract_docx_content(file_path: Path) -> Dict[str, Any]:
    """Extrae texto de archivos DOCX y devuelve una pista de sección."""
    try:
        document = Document(str(file_path))
        paragraphs = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
        text = "\n".join(paragraphs).strip()
        section = infer_section(text)
        return {"text": text, "location_hint": section or "contenido principal"}
    except Exception:
        return {"text": "", "location_hint": "sin contenido"}


def extract_excel_content(file_path: Path) -> Dict[str, Any]:
    """Convierte hojas de Excel a texto legible."""
    try:
        workbook = load_workbook(filename=str(file_path), data_only=True)
        sheet_texts: List[str] = []
        sheet_names: List[str] = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cleaned_row = [str(value).strip() if value is not None else "" for value in row]
                rows.append(" | ".join(cleaned_row))
            if rows:
                sheet_names.append(sheet.title)
                sheet_texts.append(f"Hoja: {sheet.title}\n" + "\n".join(rows))
        text = "\n\n".join(sheet_texts).strip()
        location_hint = ", ".join(sheet_names) if sheet_names else "hojas del libro"
        return {"text": text, "location_hint": location_hint}
    except Exception:
        return {"text": "", "location_hint": "sin contenido"}


def extract_csv_text(file_path: Path) -> str:
    """Lee texto de archivos CSV."""
    try:
        with file_path.open("r", encoding="utf-8", errors="ignore") as handle:
            return handle.read().strip()
    except Exception:
        return ""


def read_plain_text(file_path: Path) -> str:
    """Lee archivos de texto plano."""
    try:
        with file_path.open("r", encoding="utf-8", errors="ignore") as handle:
            return handle.read().strip()
    except Exception:
        return ""


def clean_text(text: str) -> str:
    """Limpia y normaliza el texto extraído."""
    if not text:
        return ""

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.encode("utf-8", "ignore").decode("utf-8")
    text = re.sub(r"[\u200b\u200c\u200d]", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"(?<!\n)\n(?!\n)", " ", text)
    text = re.sub(r"(?im)^\s*(Página|Page)\s*\d+\s*$", "", text)
    text = re.sub(r"(?im)^\s*\d+\s*$", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def chunk_text(text: str, max_chars: int = 1000, overlap: int = 120) -> List[str]:
    """Divide el texto en chunks manejables con solapamiento."""
    if not text:
        return []

    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) <= max_chars:
        return [normalized]

    chunks: List[str] = []
    start = 0
    while start < len(normalized):
        end = start + max_chars
        if end < len(normalized):
            split_pos = normalized.rfind(".", start, end)
            if split_pos == -1 or split_pos <= start:
                split_pos = end
            end = split_pos + 1
        chunk = normalized[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(normalized):
            break
        start = max(start + 1, end - overlap)

    return chunks


def infer_category(file_path: Path, root_dir: Path) -> str:
    """Intenta inferir la categoría a partir de la carpeta del documento."""
    try:
        relative_path = file_path.relative_to(root_dir)
        parent_name = relative_path.parts[-2] if len(relative_path.parts) > 1 else ""
    except Exception:
        parent_name = ""

    category_map = {
        "financiero": "Financiero",
        "financiero/": "Financiero",
        "logística": "Logística",
        "logistica": "Logística",
        "responsables": "Responsables",
        "servicio al cliente": "Servicio al cliente",
        "servicio-al-cliente": "Servicio al cliente",
    }
    return category_map.get(parent_name.lower(), parent_name or "General")


def infer_section(text: str) -> str:
    """Intenta encontrar un título o sección en el texto."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines[:12]:
        if len(line) <= 120 and re.search(r"(Capítulo|Sección|Objetivo|Proceso|Política|Procedimiento|FAQ|Índice|Hoja)", line, re.IGNORECASE):
            return line
    for line in lines[:8]:
        if len(line) <= 120:
            return line
    return "contenido principal"


def extract_file_metadata(file_path: Path, root_dir: Path, raw_text: str, location_hint: str) -> Dict[str, Any]:
    """Construye los metadatos del documento para los chunks."""
    try:
        stat = file_path.stat()
        created_at = datetime.fromtimestamp(stat.st_ctime).isoformat()
        modified_at = datetime.fromtimestamp(stat.st_mtime).isoformat()
    except Exception:
        created_at = None
        modified_at = None

    metadata: Dict[str, Any] = {
        "category": infer_category(file_path, root_dir),
        "source_file": file_path.name,
        "source_path": file_path.relative_to(root_dir).as_posix() if file_path.is_absolute() else file_path.as_posix(),
        "created_at": created_at,
        "modified_at": modified_at,
        "author": None,
        "section": infer_section(raw_text),
        "location": location_hint,
        "location_type": "page_range" if file_path.suffix.lower() == ".pdf" else "section",
    }

    if file_path.suffix.lower() == ".pdf":
        try:
            reader = PdfReader(str(file_path))
            pdf_meta = reader.metadata or {}
            if getattr(pdf_meta, "author", None):
                metadata["author"] = str(getattr(pdf_meta, "author"))
            elif pdf_meta.get("/Author"):
                metadata["author"] = str(pdf_meta.get("/Author"))
            if getattr(pdf_meta, "title", None):
                metadata["title"] = str(getattr(pdf_meta, "title"))
            elif pdf_meta.get("/Title"):
                metadata["title"] = str(pdf_meta.get("/Title"))
        except Exception:
            pass

    return metadata


def build_chunk_records(chunks: List[str], document_metadata: Dict[str, Any], file_path: Path) -> List[Dict[str, Any]]:
    """Asocia cada chunk con metadatos útiles para indexación y citas."""
    records: List[Dict[str, Any]] = []
    for index, chunk in enumerate(chunks, start=1):
        records.append(
            {
                "chunk_id": f"{file_path.stem}_chunk_{index}",
                "text": chunk,
                "metadata": {
                    **document_metadata,
                    "chunk_index": index,
                    "chunk_char_count": len(chunk),
                },
            }
        )
    return records


def process_documents(root_dir: Path) -> List[Dict[str, Any]]:
    """Procesa todos los documentos de una carpeta y devuelve un resumen con chunks y metadatos."""
    results: List[Dict[str, Any]] = []
    files = sorted([p for p in root_dir.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS])

    for file_path in files:
        document_content = extract_document_content(file_path)
        raw_text = document_content["text"]
        cleaned_text = clean_text(raw_text)
        chunks = chunk_text(cleaned_text)

        relative_path = file_path.relative_to(root_dir).as_posix()
        document_metadata = extract_file_metadata(file_path, root_dir, cleaned_text, document_content.get("location_hint", "contenido principal"))
        chunk_records = build_chunk_records(chunks, document_metadata, file_path)

        result = {
            "file_name": file_path.name,
            "relative_path": relative_path,
            "extension": file_path.suffix.lower(),
            "raw_chars": len(raw_text),
            "clean_chars": len(cleaned_text),
            "chunk_count": len(chunks),
            "metadata": document_metadata,
            "preview": cleaned_text[:1800],
            "chunks": [record["text"] for record in chunk_records],
            "chunk_records": chunk_records,
        }
        results.append(result)

    return results


def print_extraction_report(results: List[Dict[str, Any]], preview_chars: int = 1200) -> None:
    """Muestra una vista temporal de la extracción de cada documento."""
    print("\n=== Extracción de documentos ===")
    for index, item in enumerate(results, start=1):
        metadata = item.get("metadata", {})
        print(f"\n[{index}] {item['file_name']}")
        print(f"Ruta: {item['relative_path']}")
        print(f"Tipo: {item['extension']}")
        print(f"Categoría: {metadata.get('category', 'General')}")
        print(f"Sección: {metadata.get('section', 'contenido principal')}")
        print(f"Ubicación: {metadata.get('location', 'sin ubicación')}")
        print(f"Autor: {metadata.get('author') or 'No disponible'}")
        print(f"Caracteres originales: {item['raw_chars']}")
        print(f"Caracteres limpios: {item['clean_chars']}")
        print(f"Chunks generados: {item['chunk_count']}")
        preview = item["preview"]
        if len(preview) > preview_chars:
            preview = preview[:preview_chars] + "..."
        print("Previsualización:\n")
        print(preview)
        print("\n" + "-" * 90)


def save_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Devuelve un resumen JSON con los documentos procesados y sus metadatos."""
    return {
        "generated_at": datetime.now().isoformat(),
        "documents": results,
    }


if __name__ == "__main__":
    base_dir = Path(__file__).resolve().parent.parent
    documents_dir = base_dir / "Documentos Nexus"

    if not documents_dir.exists():
        raise FileNotFoundError(f"No se encontró la carpeta de documentos: {documents_dir}")

    results = process_documents(documents_dir)
    print_extraction_report(results)
    summary = save_summary(results)

    print(f"\nArchivos procesados: {len(results)}")
    print("Resultados disponibles en memoria para usar en el agente o para serializarlos si lo necesitas.")
    print(json.dumps(summary, ensure_ascii=False, indent=2)[:3000])
